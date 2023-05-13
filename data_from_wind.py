from WindPy import w
import openpyxl


def fetch_data_all(date: str, field: str):
    ws_para = openpyxl.load_workbook('wind参数.xlsx')['数据集参数']
    ls_place = [ws_para.cell(row=row, column=1).value for row in range(2, ws_para.max_row+1)]
    ls_code = [ws_para.cell(row=row, column=2).value for row in range(2, ws_para.max_row+1)]
    ls_cate_1 = [ws_para.cell(row=row, column=3).value for row in range(2, ws_para.max_row+1)]
    ls_cate_2 = [ws_para.cell(row=row, column=4).value for row in range(2, ws_para.max_row+1)]
    ls_cate_3 = [ws_para.cell(row=row, column=5).value for row in range(2, ws_para.max_row+1)]
    ls_cate_4 = [ws_para.cell(row=row, column=6).value for row in range(2, ws_para.max_row+1)]
    ls_cate_5 = [ws_para.cell(row=row, column=7).value for row in range(2, ws_para.max_row+1)]
    ls_para = [ls_place, ls_code, ls_cate_1, ls_cate_2, ls_cate_3, ls_cate_3, ls_cate_4, ls_cate_5]
    for ls in ls_para[2:]:
        if field in ls:
            ls_code_ = [ls_code[ind] for ind in range(len(ls_code)) if ls[ind] == field]
    # code = dict_area_code[field]
    print(ls_code_)
    w.start()
    w.isconnected()
    list_code_extracted_all = []
    list_name_extracted_all = []
    for code in ls_code_:
        print(code)

        if w.wset("sectorconstituent", "date={0};sectorid={1}".format(date, code)).Data:
            list_code_extracted = w.wset("sectorconstituent", "date={0};sectorid={1}".format(date, code)).Data[1]
            list_code_extracted_all.extend(list_code_extracted)
        if w.wset("sectorconstituent", "date={0};sectorid={1}".format(date, code)).Data:
            list_name_extracted = w.wset("sectorconstituent", "date={0};sectorid={1}".format(date, code)).Data[2]
            list_name_extracted_all.extend(list_name_extracted)

    w.stop()
    return [list_code_extracted_all, list_name_extracted_all]


# a = fetch_data_all("2023-01-30", '股票')
# b = fetch_data_all("2023-01-30", '港股')
# c = fetch_data_all("2023-01-30", '基金')
d = fetch_data_all("2023-01-30", '债券')
e = fetch_data_all("2023-01-30", '银行间')

