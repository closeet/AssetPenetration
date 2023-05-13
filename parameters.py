import openpyxl
from data_processing_func import ws_cell


def val_table_structure(path):
    wb_in = openpyxl.load_workbook(path)
    ws_in = wb_in['管理人与估值表模式']
    dict_val_struc = {}
    for row in range(2, ws_in.max_row+1):
        if ws_cell(ws_in, row, 1) is not None:
            ls_content = [ws_cell(ws_in, row, col) for col in range(2, ws_in.max_row+1) if ws_cell(ws_in, row, col) is not None]
            for col in range(2, ws_in.max_row+1):
                if ws_cell(ws_in, row, col) is None:
                    col_end = col
                    break
            ls_ind = [ws_cell(ws_in, row+1, col) for col in range(8, col_end)]
            ls_col_ind = [ls_content[6:], ls_ind]
            dict_val_struc[ws_cell(ws_in, row, 1)] = [*ls_content[:6], ls_col_ind]
    return dict_val_struc


def value_para(path):
    # wb_in = openpyxl.load_workbook('估值表参数.xlsx')
    wb_in = openpyxl.load_workbook(path)
    ls_book = []
    for ind_val_para, worksheet in enumerate(wb_in.worksheets):
        ws_in = wb_in.worksheets[ind_val_para]
        row_max = ws_in.max_row
        col_max = ws_in.max_column
        for col_val_para in range(2, col_max + 1):
            ls_col = []
            if ws_in.cell(row=1, column=col_val_para).value is not None and ws_in.cell(row=1, column=col_val_para).value != " ":
                for row in range(1, row_max):
                    value = ws_in.cell(row=row, column=col_val_para).value
                    ls_col .append(value)
                ls_book.append(ls_col)
    return ls_book


def value_name(path):
    wb_in = openpyxl.load_workbook(path, data_only=True)
    ws_in = wb_in.worksheets[0]
    ls_val = []
    for col in range(1, ws_in.max_column+1):
        ls_col = []
        for row in range(2, ws_in.max_row+1):
            value = ws_in.cell(row=row, column=col).value
            ls_col.append(value)
        ls_val.append(ls_col)
    return ls_val


ls_val_struc = val_table_structure('估值表信息.xlsx')
ls_col_ind = list(range(2, 10))
wb_para = openpyxl.load_workbook('穿透参数.xlsx')
ls_val_para = value_para('估值表参数.xlsx')
ls_val_info = value_name('估值表账面信息.xlsx')
dict_asset_type = {ws_cell(wb_para['资产类型'], row, 1): ws_cell(wb_para['资产类型'], row, 2) for row in range(2, wb_para['资产类型'].max_row+1)}
ls_asset_type_st = list(dict_asset_type.values())
ls_asset_name_type = [ws_cell(wb_para['资产类型'], row, 3) for row in range(2, wb_para['资产类型'].max_row+1)]
dict_name_type = {}
for ind, type_st in enumerate(ls_asset_type_st):
    if type_st not in dict_name_type.keys():
        dict_name_type[type_st] = ls_asset_name_type[ind]
ls_para_asset_type = dict_asset_type.keys()
ls_para_asset_type_interest = [ws_cell(wb_para['资产类型'], row, 1) for row in range(2, wb_para['资产类型'].max_row+1) if ws_cell(wb_para['资产类型'], row, 4) == 1]