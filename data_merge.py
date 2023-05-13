import openpyxl
from file_transformation_and_extraction import extract_file


def merge_file(file_path):
    # file_path = 'source/202212/金市/大家1230/大家1230填写'
    ls_file_path = extract_file(file_path)
    wb_result_template = openpyxl.load_workbook('穿透结果模板.xlsx')
    ws_result_template = wb_result_template.worksheets[0]
    path_merged_result = '/'.join(file_path.split('/')[:-1]) + '/汇总结果.xlsx'

    for path in ls_file_path:
        ws_file = openpyxl.load_workbook(path).worksheets[0]
        start_row = 1
        start_col = 1
        for row in range(1, ws_file.max_row+1):
            for col in range(1, ws_file.max_column+1):
                if ws_file.cell(row=row, column=col).value == "资产简称":
                    start_row = row
                    start_col = col
        end_row = ws_file.max_row
        for row in range(start_row, ws_file.max_row+2):
            if ws_file.cell(row=row, column=start_col).value is None:
                end_row = row
                break
        end_column = ws_file.max_column
        for column in range(start_col, ws_file.max_column+2):
            if ws_file.cell(row=start_row, column=column).value is None:
                end_column = column
                break
        start_row_result = ws_result_template.max_row
        for row in range(1, ws_result_template.max_row+2):
            if ws_result_template.cell(row=row, column=1).value is None:
                start_row_result = row
                break
        for row in range(end_row-start_row-1):
            for col in range(end_column-start_col):
                ws_result_template.cell(row=start_row_result+row, column=1+col).value = ws_file.cell(row=row+start_row+1, column=col+start_col).value
    wb_result_template.save(path_merged_result)


# merge_file('source/202212/金市/大家1231填写/1231填写')
# merge_file('source/202212/固收/固收')
merge_file('source/202212/权益/权益')