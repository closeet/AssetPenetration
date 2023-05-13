import os
import re
import time
import calendar
import openpyxl
import datetime
import pandas
from file_transformation_and_extraction import extract_file
from parameters import *
from data_processing_func import *


def extract_from_data_with_code(ls_code_to_extract: list, dict_valuation_table: dict, ls_valuation_code_parental: list,
                                dict_para_asset_type: dict, ls_code_exempt_forward: list = []) -> list:
    dict_info = {}
    dict_code_name = dict(zip(dict_valuation_table['科目代码'], dict_valuation_table['科目名称']))
    for code_to_extract in [*ls_code_to_extract, *ls_code_exempt_forward]:
        ind = dict_valuation_table['科目代码'].index(code_to_extract)
        dict_info[code_to_extract] = [dict_valuation_table[ind_val_table][ind] for ind_val_table in dict_valuation_table.keys()]
        path = [parent_code for parent_code in ls_valuation_code_parental if
                re.match(re.compile('^' + parent_code + '\\w*'), code_to_extract)]
        path.append(code_to_extract)
        path_name = [dict_code_name[code] for code in path]
        asset_type = '资产类别：未匹配'
        for code_name in reversed(path_name):
            for standard_type in dict_para_asset_type.keys():
                if standard_type in code_name:
                    asset_type = dict_para_asset_type[standard_type]
                    break
        if asset_type in ['债券', '股票', '金融期货', '基金']:
            standard = '标准'
        else:
            standard = '非标'
        dict_info[code_to_extract].append(standard)
        dict_info[code_to_extract].append(path)
        dict_info[code_to_extract].append(path_name)
        dict_info[code_to_extract].append(asset_type)
        if code_to_extract in ls_code_exempt_forward:
            dict_info[code_to_extract][5] = abs(dict_info[code_to_extract][5])
            dict_info[code_to_extract][8] = abs(dict_info[code_to_extract][8])
            dict_info[code_to_extract][10] = abs(dict_info[code_to_extract][10])
        dict_info = {code_: dict_info[code_] for code_ in dict_info.keys() if dict_info[code_][8] != 0}
    return dict_info


# 提取估值表中的所有明细信息以及设定的总体信息：总资产、总负债、总净值、总份额
def all_data_from_val(ws_valuation: openpyxl.workbook.Workbook.worksheets, dict_struc: dict) -> object:
    # 依靠管理人类型获取对应的估值表结构
    ls_struc = dict_struc[manager]
    # 起点识别
    start_place = ls_struc[0]
    # 终点识别
    end_place = ls_struc[1]
    # 列名结构识别
    col_struc = ls_struc[6]
    # 总资产识别
    total_asset_place = ls_struc[2]
    # 总负债识别
    total_liab_place = ls_struc[3]
    # 总净值识别
    total_equity_place = ls_struc[4]
    # 总份额识别
    total_share = ls_struc[5]

    # 资产负债类识别对应行坐标
    for row in range(1, ws_valuation.max_row + 1):
        if ws_cell(ws_valuation, row, 1) == start_place:
            row_start_raw = row + 1
        if ws_cell(ws_valuation, row, 1) == end_place:
            row_end_raw = row
        if ws_cell(ws_valuation, row, 1) == total_asset_place:
            row_total_asset = row
        if ws_cell(ws_valuation, row, 1) == total_liab_place:
            row_total_liab = row
        if ws_cell(ws_valuation, row, 1) == total_equity_place:
            row_total_equity = row
        if ws_cell(ws_valuation, row, 1) == total_share:
            row_total_share = row
    # 起点行坐标
    for row in range(row_start_raw, row_end_raw):
        if ws_cell(ws_valuation, row, 1) is not None:
            row_start = row
            break
    # 终点行坐标
    for row in reversed(range(row_start_raw, row_end_raw)):
        if ws_cell(ws_valuation, row, 1) is not None:
            row_end = row + 1
            break

    # 识所有别有效行坐标
    ls_row_num = [row for row in range(row_start, row_end) if ws_cell(ws_valuation, row, 1) is not None and ws_cell(ws_valuation, row, 1) != "" and ws_cell(ws_valuation, row, 1) != " "]
    # 识别所有有效列名称
    ls_col_to_extract = [col for ind, col in enumerate(col_struc[0]) if col_struc[1][ind] is not None]
    # 识别所有数值类型列坐标
    ls_keys_num = ls_col_to_extract[2:]

    # 初始化估值表所有信息数据字典
    dict_valuation_table_all_data = {'index': ls_row_num}
    # 初始化估值表总体资产负债信息列表
    ls_general_info = []

    # 提取程序主体
    # 对所有有效行列坐标进行循环并转换对应数据类型
    for col in range(1, ws_valuation.max_column + 1):
        col_name = ws_cell(ws_valuation, row_start_raw - 1, col)
        ls_col_content = []
        if col_name in ls_col_to_extract:
            # 循环提取数据
            for row in ls_row_num:
                data = ws_cell(ws_valuation, row, col)
                # 数据格式处理
                if col == 1:
                    if data is not None and isinstance(data, int):
                        data = str(ws_cell(ws_valuation, row, col))
                if col_name in ls_keys_num:
                    if data is not None and isinstance(data, str):
                        if re.match(re.compile(r'-?\d+'), data):
                            # print(row, col, data)
                            if data.find("%") + 1:
                                data = float(ws_cell(ws_valuation, row, col).replace(',', '').replace('%', "")) / 100
                            else:
                                data = float(ws_cell(ws_valuation, row, col).replace(',', ''))
                        elif data == '':
                            data = 0
                        else:
                            0
                            # print('exception')
                else:
                    0
                ls_col_content.append(data)
            dict_valuation_table_all_data[col_name] = ls_col_content
        if col_name == ls_col_to_extract[7]:
            if manager != "财务":
                for row in [row_total_asset, row_total_liab, row_total_equity, row_total_share]:
                    cell_value = ws_cell(ws_valuation, row, col)
                    if isinstance(cell_value, str):
                        total_value = float(cell_value.replace(',', '').replace('，', ''))
                    else:
                        total_value = float(ws_cell(ws_valuation, row, col))
                    ls_general_info.append(total_value)
            else:
                for row in [row_total_asset, row_total_liab, row_total_equity]:
                    cell_value = ws_cell(ws_valuation, row, col)
                    if isinstance(cell_value, str):
                        total_value = float(cell_value.replace(',', '').replace('，', ''))
                    else:
                        total_value = float(ws_cell(ws_valuation, row, col))
                    ls_general_info.append(total_value)
                ls_general_info.append(0)

    return [dict_valuation_table_all_data, ls_general_info]


def match_attributes_with_asset_raw(dict_data_to_match, dict_asset_info):
    dict_principal_interest = {}
    dict_principal_interest_code = {}

    for code_interest in dict_data_to_match.keys():
        interest_path = dict_data_to_match[code_interest][13]
        interest_type = dict_data_to_match[code_interest][14]
        if interest_type == '债券':
            for code in [code for code in dict_asset_info.keys() if code[:4] == code_interest[:4]]:
                if dict_data_to_match[code_interest][2] == dict_asset_info[code][2]:
                    # print(code_interest, code, '债券')
                    dict_principal_interest[code] = dict_data_to_match[code_interest][8]
                    dict_principal_interest_code[code] = code_interest
                    continue

        else:
            ls_code_to_match = []
            ls_interest_code_to_match = []
            for code in [code for code in dict_asset_info.keys() if code[:4] == code_interest[:4]]:
                asset_type = dict_asset_info[code][14]
                # print([code_interest, code])
                # print(interest_type, asset_type)
                if asset_type == interest_type:
                    ls_code_to_match.append(code)
                    ls_interest_code_to_match.append(code_interest)
            # print("ls_code_to_match:{0}".format(ls_code_to_match))
            if ls_code_to_match:
                0
            else:
                for code in dict_asset_info.keys():
                    asset_type = dict_asset_info[code][14]
                    if asset_type == interest_type:
                        ls_code_to_match.append(code)
                        ls_interest_code_to_match.append(code_interest)
            if len(ls_code_to_match) == 1 and ls_code_to_match[0] not in dict_principal_interest.keys():
                # print(code_interest, ls_code_to_match[0], '类型唯一对应')
                dict_principal_interest[ls_code_to_match[0]] = dict_data_to_match[code_interest][8]
                dict_principal_interest_code[ls_code_to_match[0]] = code_interest
                continue

            else:
                for code in ls_code_to_match:
                    asset_path = dict_asset_info[code][13]
                    # print(interest_path[-1], asset_path[-1])
                    if interest_path[-1] == asset_path[-1]:
                        # print(code_interest, code, '终点唯一对应')
                        dict_principal_interest[code] = dict_data_to_match[code_interest][8]
                        dict_principal_interest_code[code] = code_interest
                        continue
                    else:
                        ls_trading_place = []
                        # print(identify_element_from_path(interest_path))

                        # print(identify_element_from_path(asset_path))
                        if identify_element_from_path(interest_path) is not None and identify_element_from_path(asset_path) \
                                and identify_element_from_path(interest_path) == identify_element_from_path(asset_path):
                            # print(code_interest, code, '循迹对应')
                            dict_principal_interest[code] = dict_data_to_match[code_interest][8]
                            dict_principal_interest_code[code] = code_interest
                            continue
                        else:
                            # print([code_interest, code])
                            # print([interest_path, asset_path])
                            # print([identify_element_from_path(interest_path), identify_element_from_path(asset_path)])
                            0
    return [dict_principal_interest, dict_principal_interest_code]


def match_attributes_with_asset(dict_data_to_match, dict_asset_info):
    match_attributes_with_asset_raw(dict_data_to_match, dict_asset_info)
    dict_data_to_match_ad = dict_data_to_match
    dict_asset_info_ad = dict_asset_info
    ls_unmatched_element = []
    for code_interest in dict_data_to_match.keys():
        if code_interest not in list(match_attributes_with_asset_raw(dict_data_to_match, dict_asset_info)[1].values()):
            print(code_interest, '未匹配')
            ls_unmatched_element.append(code_interest)
    for code in ls_unmatched_element:
        dict_asset_info_ad[code+'未匹配利息'] = dict_data_to_match_ad[code]
        dict_asset_info_ad[code + '未匹配利息'][2] += '未匹配'
        dict_data_to_match_ad.pop(code)
    result = match_attributes_with_asset_raw(dict_data_to_match_ad, dict_asset_info_ad)
    dict_principal_interest = result[0]
    dict_principal_interest_code = result[1]
    return [dict_principal_interest, dict_principal_interest_code, dict_asset_info_ad, dict_data_to_match_ad]


def ls_exception_code(ls_name: list, ls_code: list, ls_exception: list = None):
    # print(ls_exception)
    ls_code_exception = []
    if ls_exception is not None:
        for ind in range(0, len(ls_name)):
            for exception in ls_exception:
                if ls_name[ind].find(exception) + 1:
                    ls_code_exception.append(ls_code[ind])
    return ls_code_exception


def val_table_info():
    wb_val_table_info = openpyxl.load_workbook('估值表参数.xlsx')
    ls_val_table_info = []
    for sheet in wb_val_table_info.sheetnames:
        ws_val_table_info = wb_val_table_info[sheet]
        for col in range(1, ws_val_table_info.max_column + 1):
            ls_col = []
            for row in range(1, ws_val_table_info.max_row + 1):
                cell = ws_val_table_info.cell(row=row, column=col).value
                # print(cell)
                ls_col.append(cell)
            if ls_col not in ls_val_table_info and ls_col[0] != '产品名' and ls_col[0] is not None and ls_col[0] != '' \
                    and ls_col[0] != '':
                ls_val_table_info.append(ls_col)
    return ls_val_table_info


def val_date(year=None, month=None):
    if year is None or month is None:
        # 获取当天日期
        today = datetime.date.today()
        # 计算上一个月份的最后一天
        last_day_of_prev_month = calendar.monthrange(today.year, today.month - 1)[1]
        # 计算上一个月底日期
        if today.month == 1:
            prev_month_end = today.replace(year=today.year - 1, month=12, day=last_day_of_prev_month)
        else:
            prev_month_end = today.replace(month=today.month - 1, day=last_day_of_prev_month)
    else:
        last_day_of_prev_month = calendar.monthrange(year, month)[1]
        prev_month_end = datetime.date(year, month, last_day_of_prev_month)
    valuation_date_func = 'DATE({0},{1},{2})'.format(prev_month_end.year, prev_month_end.month, prev_month_end.day)
    return valuation_date_func


# 从科目代码列表中分离底层科目与父级科目
def seperate_subject(ls_code_valuation):
    # 初始化父子科目列表
    ls_code_son = []
    ls_code_father = []
    # 对所有科目进行循环--1
    for index, code in enumerate(ls_code_valuation):
        # 对当前循环科目代码之后所有科目代码进行遍历--2
        for id_check in range(index + 1, len(ls_code_valuation)):
            # 构建正则检查当前循环科目1是否处于处于其后续科目2的前几位中
            if re.match(re.compile("^{0}\w*".format(code)), ls_code_valuation[id_check]):
                # 若一个科目在其后续科目中的首位被检出，添加为父级科目
                if code not in ls_code_father:
                    ls_code_father.append(code)
    # 分离子级科目
    for code in ls_code_valuation:
        # 将所有科目中的父级科目排除后即为子级科目
        if code not in ls_code_father:
            ls_code_son.append(code)
    dict_father_son_code = {'father': ls_code_father, 'son': ls_code_son}
    return dict_father_son_code


time_start = time.time()

initial_path = 'source/test'
ls_file_holder_path = extract_file('source/test')
result_all_path: str = '/'.join(initial_path.split('/')[:-1])+'/穿透结果汇总1.xlsx'
if os.path.exists(result_all_path):
    os.remove(result_all_path)

manager = '财务'
valuation_date = val_date()

for file_path in ls_file_holder_path:
    # 读取估值表
    wb_valuation = openpyxl.load_workbook(file_path)
    ws_valuation = wb_valuation.worksheets[0]
    # 从路径中分离产品名称
    product_name = ".".join(file_path.split("/")[-1].split(".")[:-1])
    # 循环表征 滚动显示提取的估值表名
    print(product_name)
    # 构建结果存储路径
    result_path: str = '/'.join(file_path.split('/')[:-2]) + '/{0}穿透结果/'.format(file_path.split('/')[-2])
    # 使用all_data_from_val 提取估值表中的总资产、负债、净值
    ls_general_info_val = all_data_from_val(ws_valuation, ls_val_struc)[1]
    print(ls_general_info_val)
    # 使用all_data_from_val 提取估值表中所有的科目数据
    dict_valuation_table = all_data_from_val(ws_valuation, ls_val_struc)[0]
    # 构建科目代码与科目名称字典
    dict_code_name = dict(zip(dict_valuation_table['科目代码'], dict_valuation_table['科目名称']))
    # 构建科目代码列表
    ls_code_valuation = dict_valuation_table['科目代码']
    # 构建科目名称列表
    ls_name_valuation = dict_valuation_table['科目名称']
    # 构建市值列表
    ls_price_valuation = list(dict_valuation_table['市值'])

    ls_code_son = seperate_subject(ls_code_valuation)['son']
    ls_code_father = seperate_subject(ls_code_valuation)['father']



    # 将当前分离的子级科目拉平，将利息、股利、减值科目横向合并，对负债、公允价值科目进行纵向分离，将清算款、期货进行分离后再纵向并入
    # 对所有需要操作的科目进行分离准备
    ls_drop_fair_value = []
    ls_drop_liab = [code for code in ls_code_son if code[0] == "2"]
    ls_drop_negative_settlement = []
    ls_drop_forward = []
    ls_forward_initial_value = []
    ls_drop_interest_all = []
    ls_drop_depreciation = []
    # 对需要处理的父级科目进行识别和记录 使用ls_exception_code函数 对路径进行回溯检测查找特定关键字
    ls_code_fair_value = ls_exception_code(ls_name_valuation, ls_code_valuation, ['公允价值变动'])
    ls_code_settlement = ls_exception_code(ls_name_valuation, ls_code_valuation, ['清算资金往来', '证券清算款'])
    ls_code_forward = ls_exception_code(ls_name_valuation, ls_code_valuation, ['套期工具'])
    ls_code_forward_initial_value = [ls_code_valuation[ind] for ind in range(0, len(ls_name_valuation)) if
                                     re.match(re.compile(r'^初始合约价值\w*'), ls_name_valuation[ind])]
    ls_code_interest = ls_exception_code(ls_name_valuation, ls_code_valuation, ['利息'])
    ls_code_depreciation = [ls_code_valuation[ind] for ind in range(0, len(ls_name_valuation)) if
                        ls_name_valuation[ind].find('减值') + 1]
    # 大家的估值表中并未在估值表路径中对利息和本金进行充分的区分，其中1002.03.03科目默认记为利息
    if manager == '大家':
        ls_code_interest.append('1002.03.03')

    # 将所有需要处理的父级科目与子级科目进行匹配，将特定父级科目路径下的所有子级科目进行分离
    for code in ls_code_son:
        for code_fair_value in ls_code_fair_value:
            if re.match(re.compile('^' + code_fair_value + '\\w*'), code):
                if code not in ls_drop_fair_value:
                    ls_drop_fair_value.append(code)
        for code_settlement in ls_code_settlement:
            if re.match(re.compile('^' + code_settlement + '\\w*'), code):
                if ls_price_valuation[ls_code_valuation.index(code)] < 0:
                    if code not in ls_drop_negative_settlement:
                        ls_drop_negative_settlement.append(code)
        for code_forward in ls_code_forward:
            if re.match(re.compile('^' + code_forward + '\\w*'), code):
                if code not in ls_drop_forward:
                    ls_drop_forward.append(code)
                for code_initial_value in ls_code_forward_initial_value:
                    if re.match(re.compile('^' + code_initial_value + '\\w*'), code):
                        if code not in ls_forward_initial_value:
                            ls_forward_initial_value.append(code)
        for code_interest in ls_code_interest:
            if re.match(re.compile('^' + code_interest + '\\w*'), code):
                if code not in ls_drop_interest_all:
                    ls_drop_interest_all.append(code)
        for code_depreciation in ls_code_depreciation:
            if re.match(re.compile('^' + code_depreciation + '\\w*'), code):
                if code not in ls_drop_depreciation:
                    ls_drop_depreciation.append(code)
    # 将所有的非本金科目加以汇总
    ls_drop = [*ls_drop_liab, *ls_drop_negative_settlement, *ls_drop_fair_value, *ls_drop_forward,
               *ls_drop_interest_all, *ls_drop_depreciation]
    # 对非本金科目去除后得到子级本金科目
    ls_asset_code = [code for code in ls_code_son if code not in ls_drop]
    # 获取所有的子级利息科目
    ls_interest = [code for code in ls_drop_interest_all if code[0] == '1']
    # 获取所有的子级减值科目
    ls_depreciation = [code for code in ls_drop_depreciation if code[0] == '1']
    # 将期货科目从完整数据中提取出其他信息补完为完整资产字典
    dict_asset_info = extract_from_data_with_code(ls_asset_code, dict_valuation_table, ls_code_father,
                                                  dict_asset_type, ls_forward_initial_value)
    # 将利息科目从完整数据中提取出其他信息补完为完整资产字典
    dict_interest_info = extract_from_data_with_code(ls_interest, dict_valuation_table, ls_code_father,
                                                     dict_asset_type)
    # 将减值科目从完整数据中提取出其他信息补完为完整资产字典
    dict_depreciation_info = extract_from_data_with_code(ls_depreciation, dict_valuation_table, ls_code_father,
                                                     dict_asset_type)
    # 将利息与本金进行匹配
    result_match_interest = match_attributes_with_asset(dict_interest_info, dict_asset_info)
    # 生成利息本金数值匹配字典在过程中用以检查匹配结果
    dict_principal_interest = result_match_interest[0]
    # 生成利息本金代码匹配字典在过程中用以检查匹配结果
    dict_principal_interest_code = result_match_interest[1]
    # 为未匹配的利息生成一笔对应的本金调整后的本金字典
    dict_asset_info = result_match_interest[2]
    # 将合并利息的本金字典匹配减值信息
    result_match_depreciation = match_attributes_with_asset(dict_depreciation_info, dict_asset_info)
    # 生成减值本金数值匹配字典在过程中用以检查匹配结果
    dict_principal_depreciation = result_match_depreciation[0]
    # 生成减值本金代码匹配字典在过程中用以检查匹配结果
    dict_principal_depreciation_code = result_match_depreciation[1]
    # 为未匹配的利息生成一笔对应的本金调整后的本金字典
    dict_asset_info = result_match_depreciation[2]

    for code in dict_asset_info.keys():

        if code in dict_principal_interest:
            interest = dict_principal_interest[code]
        else:
            interest = 0
        if code in dict_principal_depreciation:
            dict_asset_info[code][5] = dict_asset_info[code][5] + dict_principal_depreciation[code]
            dict_asset_info[code][8] = dict_asset_info[code][8] + dict_principal_depreciation[code]
        if dict_asset_info[code][11] == '标准':
            asset_abbr = dict_asset_info[code][2]
        elif dict_asset_info[code][14] == '保险资管产品':
            asset_abbr = dict_asset_info[code][2]
        elif dict_name_type[dict_asset_info[code][14]] == "存款":
            bank = identify_bank_element_from_path(dict_asset_info[code][13])
            asset_abbr = bank + dict_asset_info[code][14]
        elif dict_name_type[dict_asset_info[code][14]] == "保证备付金":
            element = identify_element_from_path(dict_asset_info[code][13])
            asset_abbr = dict_asset_info[code][14] + '-' + element
        else:
            used_element = []
            for subject in reversed(dict_asset_info[code][13]):
                if subject == dict_asset_info[code][14]:
                    used_element.append(subject)
                    break
                else:
                    used_element.append(subject)
            asset_abbr = '-'.join(reversed(used_element))
        dict_asset_info[code].append(interest)
        dict_asset_info[code].append(asset_abbr)
        dict_asset_info[code].append(product_name)
    ls_asset_info = [ls_asset for ls_asset in dict_asset_info.values()]
    wb_result = openpyxl.Workbook()
    ws_result = wb_result.active
    ls_title_name = ['估值表中位置', '科目代码', '底层科目名称', '数量', '单位成本', '成本', '成本占净值', '市价',
                     '市值', '市值占净值',
                     '估值增值', '公开交易', '科目路径', '名称路径', '资产类型', '利息', '资产简称',
                     '表层资产简称']
    for col in range(len(ls_title_name)):
        ws_result.cell(row=1, column=col + 1).value = ls_title_name[col]
        for row in range(len(list(dict_asset_info.keys()))):
            if col == 12 or col == 13:
                ws_result.cell(row=row + 2, column=col + 1).value = '-'.join(
                    dict_asset_info[list(dict_asset_info.keys())[row]][col])
            else:
                ws_result.cell(row=row + 2, column=col + 1).value = dict_asset_info[list(dict_asset_info.keys())[row]][
                    col]
    if not os.path.exists(result_path):
        os.mkdir(result_path)

    wb_result.save(result_path + '/{0}穿透资产数据.xlsx'.format(product_name))

    if os.path.exists(result_all_path):
        wb_result_all = openpyxl.load_workbook(result_all_path)
    else:
        wb_result_all = openpyxl.load_workbook('穿透结果模板.xlsx')
    ws_result_all = wb_result_all.worksheets[0]
    for row_num in range(1, ws_result_all.max_row + 2):
        if ws_cell(ws_result_all, row_num, 1) is None:
            start_row_all = row_num
            break
        else:
            0
    ls_col_input_result_all = [1, 7, 8, 9, 10, 67, 68, 69]
    ls_col_source_result = [16, 3, 5, 8, 15, 14, 11, 17]
    for ind_match_solv_valuation in range(len(ls_col_input_result_all)):
        for row_match_solv_val in range(len(ls_asset_info)):
            ws_result_all.cell(row_match_solv_val + start_row_all,
                               ls_col_input_result_all[ind_match_solv_valuation]).value = \
            ls_asset_info[row_match_solv_val][ls_col_source_result[ind_match_solv_valuation]]
    wb_result_all.save(result_all_path)

wb_result_all = openpyxl.load_workbook(result_all_path)
ws_result_all = wb_result_all.worksheets[0]
ls_col_all_result = list(range(1, ws_result_all.max_column+1))
ls_col_formula = [item for item in ls_col_all_result if item not in ls_col_input_result_all]
for row in range(2, ws_result_all.max_row+1):
    ws_result_all.cell(row=row, column=2).value = """=IF($BR{0}="股票",s_info_compname($E{0}),IF($BR{0}="债券",b_info_fullname($E{0}),IF($BR{0}="基金",F_Info_FullName($E{0}),$A{0})))""".format(str(row))
    ws_result_all.cell(row=row, column=3).value = """=IFERROR(VLOOKUP($D{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$F$1:$G$93,2,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=4).value = """=IFERROR(VLOOKUP(@IF(ISERROR(FIND("股指期货",$A{0})),IF(ISERROR(FIND("HK",$A{0})),IF(ISERROR(FIND("永续",$A{0})),IF($BO{0}="股票",s_info_mkt($E{0}),IF($BO{0}="债券",b_info_windl2type($E{0}),IF($BO{0}="基金",f_info_firstinvesttype($E{0}),IFERROR(VLOOKUP($BO{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$AC:$AD,2,0),"")))),"无固定期限资本债券"),"香港股票"),"股指期货空头合约"),'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$C:$D,2,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=5).value = """=IF($BP{0}="标准",IF(to_windcode($A{0})=0,to_windcode(VLOOKUP($A{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]代码识别'!$A:$B,2,0)),to_windcode($A{0})),"")""".format(str(row))
    ws_result_all.cell(row=row, column=6).value = """=@IF($BP{0}="标准",s_info_compname($E{0}),IFERROR(VLOOKUP($D{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]交易对手识别'!$E:$F,2,0),IFERROR(VLOOKUP($A{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]交易对手识别'!$A:$B,2,0),IF($D{0}="活期存款",VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[资管产品参数.xlsx]Sheet1'!$A:$E,5,0)))))""".format(str(row))
    ws_result_all.cell(row=row, column=12).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]表层参数'!$B:$C,2,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=13).value = """=VLOOKUP($BQ{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$A:$G,2,0)""".format(str(row))
    ws_result_all.cell(row=row, column=14).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$B:$M,7,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=15).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$B:$M,8,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=16).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$B:$M,9,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=17).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$B:$M,10,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=18).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]表层参数'!$B:$J,7,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=19).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$B:$M,11,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=26).value = """=@IF($C{0}="股票",s_info_compindex2($E{0},3,'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$A$2),"")""".format(str(row))
    ws_result_all.cell(row=row, column=27).value = """=IF($C{0}="股票",$I{0}/@s_val_ev($E{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$A$2,1),"")""".format(str(row))
    ws_result_all.cell(row=row, column=28).value = """=IF(OR($C{0}="优先股",$C{0}="无固定期限资本债券"),"是","")""".format(str(row))
    ws_result_all.cell(row=row, column=29).value = """=IF(OR($C{0}="优先股",$C{0}="无固定期限资本债券"),"银行","")""".format(str(row))
    ws_result_all.cell(row=row, column=30).value = """=IF($AC{0}="银行",VLOOKUP($F{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]银行法人机构'!$B:$F,4,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=31).value = """=IF($AC{0}="银行",VLOOKUP($F{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$J:$M,2,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=32).value = """=IF($AC{0}="银行",VLOOKUP($F{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$J:$M,3,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=33).value = """=IF($AC{0}="银行",VLOOKUP($F{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$J:$M,4,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=39).value = """=IF($BR{0}="金衍",$BQ{0},"")""".format(str(row))
    ws_result_all.cell(row=row, column=40).value = """=IF($AM{0}="","",VLOOKUP($A{0}&$BQ{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$T:$Y,4,0))""".format(str(row))
    ws_result_all.cell(row=row, column=41).value = """=IF($AM{0}="","",VLOOKUP($A{0}&$BQ{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$T:$Y,5,0))""".format(str(row))
    ws_result_all.cell(row=row, column=42).value = """=IF($AM{0}="","",VLOOKUP($A{0}&$BQ{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$T:$Y,6,0))""".format(str(row))
    ws_result_all.cell(row=row, column=47).value = """=IF(ISERROR(FIND("HK",$E{0})),"","香港")""".format(str(row))
    ws_result_all.cell(row=row, column=49).value = """=IF($C{0}="银行存款",VLOOKUP($F{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]银行法人机构'!$B:$F,4,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=50).value = """=IF($C{0}="银行存款",VLOOKUP($F{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$J:$K,2,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=51).value = """=IF($BR{0}="债券",b_anal_ptmyear($E{0},{1}),"")""".format(str(row), valuation_date)
    ws_result_all.cell(row=row, column=52).value = """=@IF(AND($BR{0}="债券",AND($D{0}<>"国债",$D{0}<>"政策性金融债券",$D{0}<>"地方政府债",$D{0}<>"政府支持机构债券")),IF(@b_rate_latestcredit($E{0})=0,s_rate_latestissurercreditrating($E{0}),b_rate_latestcredit($E{0})),"")""".format(str(row))
    ws_result_all.cell(row=row, column=53).value = """=@IF($BR{0}="债券",IF(@b_anal_modidura_cnbd($E{0},{1})=0,IF(@b_anal_modifiedduration($E{0},{1})=0,b_anal_ptmyear($E{0},{1}),b_anal_modifiedduration($E{0},{1})),b_anal_modidura_cnbd($E{0},{1})),"")""".format(str(row), valuation_date)
    ws_result_all.cell(row=row, column=54).value = """=IF($BR{0}="债券",IF(@b_agency_certification($E{0})=0,"否","是"),"")""".format(str(row))
    ws_result_all.cell(row=row, column=66).value = """=IFERROR(VLOOKUP($M{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[估值表账面信息.xlsx]Sheet1'!$B:$M,12,0),"")""".format(str(row))
    ws_result_all.cell(row=row, column=70).value = """=IFERROR(VLOOKUP($C{0},'C:\\Users\\closeet\\PycharmProjects\\AssetPenetration\\[穿透结果模板参数.xlsx]底层参数'!$G:$H,2,0),"")""".format(str(row))

wb_result_all.calculation_on_save = True
wb_result_all.save(result_all_path)
time_end = time.time()
print('用时{}秒'.format(time_end-time_start))